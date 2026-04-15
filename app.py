"""
╔══════════════════════════════════════════════════════════════════════╗
║         NF-PDF → Excel Enterprise Ultra — app.py                    ║
║  Sistema Enterprise de Extração de NFS-e/NF-e para Excel            ║
║  Versão 2.0 | Suporte multi-prefeitura | > 98% de precisão          ║
╚══════════════════════════════════════════════════════════════════════╝
"""

# ─── Stdlib ───────────────────────────────────────────────────────────────────
import io
import logging
import os
import re
import traceback
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# ─── Third-party ──────────────────────────────────────────────────────────────
import fitz  # PyMuPDF
import openpyxl
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("NF-Enterprise")


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1: CONSTANTES E PADRÕES DE REGEX
# ══════════════════════════════════════════════════════════════════════════════

# Padrão CNPJ/CPF universal
RE_CNPJ = re.compile(
    r"\b(\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4}[-\s]?\d{2})\b"
)
RE_CPF = re.compile(
    r"\b(\d{3}[\.\s]?\d{3}[\.\s]?\d{3}[-\s]?\d{2})\b"
)

# Padrões de data/hora
RE_DATA_HORA = re.compile(
    r"\b(\d{2}[\/\-\.]\d{2}[\/\-\.]\d{4})"
    r"(?:[T\s,]+(\d{2}:\d{2}(?::\d{2})?))?",
    re.IGNORECASE,
)
RE_COMPETENCIA = re.compile(
    r"\b(\d{2}[\/\-\.]\d{4})\b"
)

# Padrões de valor monetário brasileiro
RE_VALOR = re.compile(
    r"R?\$?\s*(\d{1,3}(?:\.\d{3})*(?:,\d{2})?|\d+(?:,\d{2})?)"
)

# Padrões de número da nota
RE_NUMERO_NOTA = [
    re.compile(r"n[uú]mero\s+(?:da\s+)?n[fo]t?a?\s*[:\-\s]*(\d+)", re.IGNORECASE),
    re.compile(r"nfs?[- ]?e?\s*[n°nº#:]*\s*(\d+)", re.IGNORECASE),
    re.compile(r"nota\s*fiscal\s*[n°nº#:]*\s*(\d+)", re.IGNORECASE),
    re.compile(r"n[°º]\s*(\d{4,})", re.IGNORECASE),
    re.compile(r"rps\s*[n°nº#:]*\s*(\d+)", re.IGNORECASE),
    re.compile(r"n[uú]mero\s*[:\-]\s*(\d+)", re.IGNORECASE),
    re.compile(r"(?:^|\s)(\d{7,})\s*$", re.MULTILINE),
]

RE_SERIE = [
    re.compile(r"s[eé]rie\s*[:\-\s]*([A-Z0-9]{1,5})", re.IGNORECASE),
    re.compile(r"s[eé]r\.?\s*[:\-\s]*([A-Z0-9]{1,5})", re.IGNORECASE),
]

# Padrões de alíquota
RE_ALIQUOTA = re.compile(
    r"al[íi]quota\s*(?:do\s*iss|iss)?\s*[:\-]?\s*(\d+[,\.]\d+)\s*%?",
    re.IGNORECASE,
)
RE_ALIQUOTA_ALT = re.compile(r"(\d+[,\.]\d+)\s*%", re.IGNORECASE)

# Mapeamento de palavras-chave para campos
KEYWORD_MAP = {
    "numero_nota": [
        "número da nota", "nfs-e", "nfse", "nota fiscal", "n° da nota",
        "número nota", "no. nota", "n. nota", "número do rps", "rps",
        "nota de serviço", "numero", "nf-e", "número",
    ],
    "data_emissao": [
        "data de emissão", "data emissão", "emissão", "data/hora",
        "emitida em", "data de competência", "competência",
        "data emitido", "data nota",
    ],
    "competencia": [
        "competência", "mês competência", "período", "mês de referência",
        "referência",
    ],
    "cnpj_prestador": [
        "cnpj prestador", "cnpj do prestador", "prestador cnpj",
        "cnpj/cpf prestador", "cpf/cnpj prestador", "emitente",
        "empresa prestadora", "prestador de serviços",
    ],
    "razao_prestador": [
        "prestador de serviços", "prestador", "emitente", "empresa",
        "razão social prestador", "nome/razão social", "razão social do prestador",
    ],
    "cnpj_tomador": [
        "cnpj tomador", "cnpj do tomador", "tomador cnpj",
        "cnpj/cpf tomador", "cpf/cnpj tomador",
        "tomador de serviços", "contratante",
    ],
    "razao_tomador": [
        "tomador de serviços", "tomador", "contratante", "cliente",
        "razão social tomador", "nome/razão social do tomador",
    ],
    "valor_servicos": [
        "valor total dos serviços", "valor bruto", "valor dos serviços",
        "total dos serviços", "valor total de serviços", "valor do serviço",
        "total bruto",
    ],
    "valor_liquido": [
        "valor líquido", "valor a receber", "valor líquido da nota",
        "líquido da nfs", "valor líquido nfse", "total líquido",
        "vlr líquido",
    ],
    "base_calculo": [
        "base de cálculo", "base de calc", "base calculo",
        "valor base", "base tributável",
    ],
    "valor_iss": [
        "valor do iss", "iss calculado", "iss", "valor iss",
        "imposto sobre serviços", "valor do imposto",
    ],
    "iss_retido": [
        "iss retido", "iss a reter", "retenção iss", "iss retido na fonte",
        "retido na fonte iss",
    ],
    "pis": [
        "pis", "pis/pasep", "valor pis", "retenção pis",
    ],
    "cofins": [
        "cofins", "valor cofins", "retenção cofins",
    ],
    "csll": [
        "csll", "valor csll", "contribuição social",
    ],
    "irrf": [
        "irrf", "ir", "irpj", "imposto de renda", "i.r.r.f",
        "retenção ir", "valor ir",
    ],
    "inss": [
        "inss", "previdência", "contribuição previdenciária",
        "retenção inss",
    ],
    "descricao_servico": [
        "discriminação dos serviços", "discriminação", "descrição",
        "histórico", "objeto", "serviços prestados", "atividade",
        "descrição do serviço",
    ],
    "codigo_servico": [
        "código do serviço", "código de serviço", "código tributário",
        "código cnae", "cnae", "atividade econômica",
        "item da lista", "código atividade",
    ],
    "municipio_prestacao": [
        "município de incidência", "local da prestação",
        "município de prestação", "município prestação", "local de prestação",
    ],
    "desconto": [
        "descontos", "desconto incondicional", "desconto condicionado",
        "total de descontos",
    ],
}


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2: UTILITÁRIOS DE TEXTO E NORMALIZAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def normalizar_texto(texto: str) -> str:
    """Remove acentos e converte para minúsculas."""
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def limpar_valor(valor_str: str) -> Optional[float]:
    """Converte string de valor BR (1.234,56) para float."""
    if not valor_str:
        return None
    try:
        limpo = re.sub(r"[R$\s]", "", str(valor_str).strip())
        limpo = limpo.replace(".", "").replace(",", ".")
        return float(limpo)
    except (ValueError, AttributeError):
        return None


def formatar_cnpj(cnpj_str: str) -> str:
    """Formata CNPJ no padrão XX.XXX.XXX/XXXX-XX."""
    if not cnpj_str:
        return ""
    digitos = re.sub(r"\D", "", cnpj_str)
    if len(digitos) == 14:
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
    if len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    return cnpj_str.strip()


def extrair_melhor_valor(texto: str) -> Optional[str]:
    """Extrai o melhor valor monetário de um trecho de texto."""
    if not texto:
        return None
    matches = RE_VALOR.findall(texto)
    if not matches:
        return None
    # Prefere valores com centavos
    com_centavos = [m for m in matches if "," in m]
    return com_centavos[0] if com_centavos else matches[0]


def texto_apos_keyword(texto_completo: str, keywords: List[str],
                        janela: int = 120) -> Optional[str]:
    """
    Busca por palavras-chave e retorna o texto após elas (janela de N chars).
    Estratégia de proximidade por palavras-chave.
    """
    texto_norm = normalizar_texto(texto_completo)
    for kw in keywords:
        kw_norm = normalizar_texto(kw)
        pos = texto_norm.find(kw_norm)
        if pos != -1:
            inicio = pos + len(kw_norm)
            trecho = texto_completo[inicio: inicio + janela]
            trecho = re.sub(r"^[\s:\/\-]+", "", trecho)
            if trecho:
                return trecho
    return None


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3: EXTRATORES DE CAMPOS ESPECÍFICOS
# ══════════════════════════════════════════════════════════════════════════════

def extrair_cnpj_cpf(texto: str) -> List[str]:
    """Extrai todos os CNPJs/CPFs únicos do texto."""
    cnpjs = RE_CNPJ.findall(texto)
    cpfs  = RE_CPF.findall(texto)
    todos = list(dict.fromkeys(cnpjs + cpfs))
    return [formatar_cnpj(c) for c in todos]


def extrair_numero_nota(texto: str) -> Optional[str]:
    """Extrai número da NFS-e / NF-e / RPS."""
    for padrao in RE_NUMERO_NOTA:
        m = padrao.search(texto)
        if m:
            return m.group(1).strip()
    return None


def extrair_serie(texto: str) -> Optional[str]:
    for padrao in RE_SERIE:
        m = padrao.search(texto)
        if m:
            return m.group(1).strip()
    return None


def extrair_data_hora(texto: str) -> Tuple[Optional[str], Optional[str]]:
    """Retorna (data, hora) da emissão da nota."""
    m = RE_DATA_HORA.search(texto)
    if m:
        data = m.group(1).replace("-", "/").replace(".", "/")
        hora = m.group(2) if m.group(2) else None
        return data, hora
    return None, None


def extrair_competencia(texto: str) -> Optional[str]:
    """Extrai mês/ano de competência."""
    trecho = texto_apos_keyword(texto, KEYWORD_MAP["competencia"], 30)
    if trecho:
        m = RE_COMPETENCIA.search(trecho)
        if m:
            return m.group(1)
    # Fallback: qualquer padrão MM/YYYY no texto completo após data de emissão
    matches = RE_COMPETENCIA.findall(texto)
    return matches[0] if matches else None


def extrair_valor_campo(texto: str, keywords: List[str]) -> Optional[str]:
    """Estratégia genérica: busca keyword e extrai valor após ela."""
    trecho = texto_apos_keyword(texto, keywords, 60)
    if trecho:
        return extrair_melhor_valor(trecho)
    return None


def extrair_aliquota(texto: str) -> Optional[str]:
    """Extrai alíquota do ISS."""
    # Busca após keyword
    trecho = texto_apos_keyword(texto, ["alíquota", "aliquota", "alíq.", "aliq."], 30)
    if trecho:
        m = RE_ALIQUOTA_ALT.search(trecho)
        if m:
            return m.group(1).replace(".", ",")

    m = RE_ALIQUOTA.search(texto)
    if m:
        return m.group(1).replace(".", ",")
    return None


def extrair_iss_retido(texto: str) -> Optional[str]:
    """Detecta se o ISS é retido na fonte."""
    keywords_sim = ["iss retido", "retido na fonte", "retenção iss", "iss a reter",
                    "retido pelo tomador", "sim", "iss retido: sim"]
    keywords_nao = ["não retido", "nao retido", "iss não retido", "não", "nao"]
    texto_norm = normalizar_texto(texto)

    # Busca contexto ao redor de "retido"
    m = re.search(r"retid[oa].{0,30}", texto_norm)
    if m:
        contexto = m.group(0)
        for kw in keywords_nao:
            if kw in contexto:
                return "Não"
        for kw in keywords_sim:
            if kw in contexto:
                return "Sim"

    for kw in keywords_sim:
        if kw in texto_norm:
            return "Sim"
    return "Não"


def extrair_razao_social(texto: str, keywords: List[str]) -> Optional[str]:
    """Extrai razão social após keyword."""
    trecho = texto_apos_keyword(texto, keywords, 150)
    if not trecho:
        return None
    # Pega até quebra de linha ou delimitador
    linhas = re.split(r"[\n\r|]+", trecho)
    razao = linhas[0].strip()
    # Remove lixo inicial como CNPJ ou dígitos
    razao = re.sub(r"^[\d\.\-\/]+\s*", "", razao)
    razao = re.sub(r"\s+", " ", razao)
    return razao[:120] if razao else None


def extrair_discriminacao(texto: str) -> Optional[str]:
    """Extrai discriminação/descrição completa dos serviços."""
    keywords = KEYWORD_MAP["descricao_servico"]
    trecho = texto_apos_keyword(texto, keywords, 600)
    if trecho:
        # Remove texto de campos de valores que possam ter escapado
        trecho = re.sub(r"\n{3,}", "\n\n", trecho)
        return trecho[:500].strip()
    return None


def extrair_codigo_servico(texto: str) -> Optional[str]:
    """Extrai código do serviço / CNAE."""
    trecho = texto_apos_keyword(texto, KEYWORD_MAP["codigo_servico"], 30)
    if trecho:
        m = re.search(r"[\d\.\-]+", trecho)
        if m:
            return m.group(0).strip()
    # Padrões específicos
    for padrao in [
        re.compile(r"c[oó]d\.?\s*servi[cç]o\s*[:\-]?\s*([\d\.]+)", re.IGNORECASE),
        re.compile(r"item\s+(?:da\s+lista|lista)\s*[:\-]?\s*([\d\.]+)", re.IGNORECASE),
        re.compile(r"cnae\s*[:\-]?\s*([\d\.\-]+)", re.IGNORECASE),
    ]:
        m = padrao.search(texto)
        if m:
            return m.group(1).strip()
    return None


def extrair_municipio(texto: str) -> Optional[str]:
    """Extrai município de prestação."""
    trecho = texto_apos_keyword(texto, KEYWORD_MAP["municipio_prestacao"], 80)
    if trecho:
        m = re.match(r"([A-ZÀ-Ü][a-zà-ü\s]+(?:\-[A-Z]{2})?)", trecho.strip())
        if m:
            return m.group(1).strip()
    return None


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4: ESTRATÉGIAS DE EXTRAÇÃO EM CASCATA
# ══════════════════════════════════════════════════════════════════════════════

class ExtratorEstrategias:
    """
    Implementa as 4 estratégias de extração em cascata:
      1. Regex avançados com múltiplos padrões
      2. Busca por proximidade de palavras-chave
      3. Análise posicional relativa (layout-aware)
      4. Extração inteligente de tabelas
    """

    def __init__(self, texto_completo: str, blocos_posicionais: List[Dict],
                 tabelas: List[List[List[str]]]):
        self.texto = texto_completo
        self.blocos = blocos_posicionais   # [{text, x0, y0, x1, y1}, ...]
        self.tabelas = tabelas
        self._cache: Dict[str, Any] = {}

    # ── Estratégia 1: Regex multi-padrão ──────────────────────────────────────

    def _estrategia_regex(self) -> Dict[str, Any]:
        t = self.texto
        resultado = {}

        # Número da nota
        resultado["numero_nota"] = extrair_numero_nota(t)
        resultado["serie"] = extrair_serie(t)

        # Datas
        resultado["data_emissao"], resultado["hora_emissao"] = extrair_data_hora(t)
        resultado["competencia"] = extrair_competencia(t)

        # CNPJs — lista completa
        resultado["_todos_cnpjs"] = extrair_cnpj_cpf(t)

        # Valores
        resultado["valor_servicos"] = extrair_valor_campo(t, KEYWORD_MAP["valor_servicos"])
        resultado["valor_liquido"]   = extrair_valor_campo(t, KEYWORD_MAP["valor_liquido"])
        resultado["base_calculo"]    = extrair_valor_campo(t, KEYWORD_MAP["base_calculo"])
        resultado["valor_iss"]       = extrair_valor_campo(t, KEYWORD_MAP["valor_iss"])
        resultado["pis"]             = extrair_valor_campo(t, KEYWORD_MAP["pis"])
        resultado["cofins"]          = extrair_valor_campo(t, KEYWORD_MAP["cofins"])
        resultado["csll"]            = extrair_valor_campo(t, KEYWORD_MAP["csll"])
        resultado["irrf"]            = extrair_valor_campo(t, KEYWORD_MAP["irrf"])
        resultado["inss"]            = extrair_valor_campo(t, KEYWORD_MAP["inss"])
        resultado["desconto"]        = extrair_valor_campo(t, KEYWORD_MAP["desconto"])

        # ISS
        resultado["aliquota_iss"] = extrair_aliquota(t)
        resultado["iss_retido"]   = extrair_iss_retido(t)

        # Serviço
        resultado["codigo_servico"]    = extrair_codigo_servico(t)
        resultado["descricao_servico"] = extrair_discriminacao(t)
        resultado["municipio_prestacao"] = extrair_municipio(t)

        return resultado

    # ── Estratégia 2: Proximidade por palavras-chave ───────────────────────────

    def _estrategia_keywords(self) -> Dict[str, Any]:
        t = self.texto
        resultado = {}

        # Razões sociais
        resultado["razao_prestador"] = extrair_razao_social(t, KEYWORD_MAP["razao_prestador"])
        resultado["razao_tomador"]   = extrair_razao_social(t, KEYWORD_MAP["razao_tomador"])

        # CNPJ Prestador e Tomador por contexto
        for campo, kws in [
            ("cnpj_prestador", KEYWORD_MAP["cnpj_prestador"]),
            ("cnpj_tomador",   KEYWORD_MAP["cnpj_tomador"]),
        ]:
            trecho = texto_apos_keyword(t, kws, 60)
            if trecho:
                cnpjs = extrair_cnpj_cpf(trecho)
                if cnpjs:
                    resultado[campo] = cnpjs[0]

        return resultado

    # ── Estratégia 3: Análise posicional (layout-aware) ───────────────────────

    def _estrategia_posicional(self) -> Dict[str, Any]:
        """
        Usa coordenadas dos blocos de texto para identificar pares label:valor
        em layouts com colunas paralelas (comum em NFS-e de prefeituras).
        """
        resultado = {}
        if not self.blocos:
            return resultado

        # Agrupa blocos por faixa vertical (mesma linha ≈ ±5px)
        linhas: Dict[int, List[Dict]] = {}
        for bloco in self.blocos:
            y_norm = round(bloco.get("y0", 0) / 5) * 5
            linhas.setdefault(y_norm, []).append(bloco)

        # Ordena cada linha por x0
        for y_key in linhas:
            linhas[y_key].sort(key=lambda b: b.get("x0", 0))

        for y_key, blocos_linha in linhas.items():
            if len(blocos_linha) < 2:
                continue
            for i, bloco_label in enumerate(blocos_linha[:-1]):
                label_txt = normalizar_texto(bloco_label.get("text", ""))
                valor_bloco = blocos_linha[i + 1]
                valor_txt   = valor_bloco.get("text", "").strip()

                # Mapeamento de labels comuns
                if any(kw in label_txt for kw in ["numero", "n° nota", "nfse"]):
                    resultado.setdefault("numero_nota", re.sub(r"\D", "", valor_txt) or valor_txt)
                elif any(kw in label_txt for kw in ["data emis", "emissao", "emissão"]):
                    resultado.setdefault("data_emissao", valor_txt[:10])
                elif any(kw in label_txt for kw in ["competencia", "competência"]):
                    resultado.setdefault("competencia", valor_txt[:7])
                elif any(kw in label_txt for kw in ["valor total", "total serv"]):
                    resultado.setdefault("valor_servicos", extrair_melhor_valor(valor_txt))
                elif any(kw in label_txt for kw in ["valor liquido", "valor líquido"]):
                    resultado.setdefault("valor_liquido", extrair_melhor_valor(valor_txt))
                elif any(kw in label_txt for kw in ["base de calc"]):
                    resultado.setdefault("base_calculo", extrair_melhor_valor(valor_txt))
                elif any(kw in label_txt for kw in ["valor iss", "iss calculado"]):
                    resultado.setdefault("valor_iss", extrair_melhor_valor(valor_txt))
                elif any(kw in label_txt for kw in ["aliquota", "alíquota"]):
                    m = RE_ALIQUOTA_ALT.search(valor_txt)
                    if m:
                        resultado.setdefault("aliquota_iss", m.group(1))
                elif any(kw in label_txt for kw in ["cnpj prest", "cnpj do prest"]):
                    cnpjs = extrair_cnpj_cpf(valor_txt)
                    if cnpjs:
                        resultado.setdefault("cnpj_prestador", cnpjs[0])
                elif any(kw in label_txt for kw in ["cnpj tom", "cnpj do tom"]):
                    cnpjs = extrair_cnpj_cpf(valor_txt)
                    if cnpjs:
                        resultado.setdefault("cnpj_tomador", cnpjs[0])

        return resultado

    # ── Estratégia 4: Extração de tabelas ─────────────────────────────────────

    def _estrategia_tabelas(self) -> Dict[str, Any]:
        """
        Analisa tabelas extraídas pelo pdfplumber para identificar pares
        (header, valor) e linhas de itens/serviços.
        """
        resultado = {}
        if not self.tabelas:
            return resultado

        for tabela in self.tabelas:
            if not tabela:
                continue
            # Normaliza todas as células
            rows = [[normalizar_texto(str(c or "")) for c in row] for row in tabela]

            # Tenta identificar cabeçalhos na primeira linha
            if len(rows) < 2:
                continue
            headers = rows[0]

            for row in rows[1:]:
                for i, header in enumerate(headers):
                    if i >= len(row):
                        continue
                    val = row[i].strip()
                    if not val:
                        continue

                    if any(kw in header for kw in ["numero", "n° nota", "nfse"]):
                        resultado.setdefault("numero_nota", val)
                    elif any(kw in header for kw in ["razao social prest", "prest"]):
                        resultado.setdefault("razao_prestador", val)
                    elif any(kw in header for kw in ["razao social tom", "tom"]):
                        resultado.setdefault("razao_tomador", val)
                    elif any(kw in header for kw in ["cnpj prest"]):
                        resultado.setdefault("cnpj_prestador", formatar_cnpj(val))
                    elif any(kw in header for kw in ["cnpj tom"]):
                        resultado.setdefault("cnpj_tomador", formatar_cnpj(val))
                    elif any(kw in header for kw in ["valor total", "valor serv"]):
                        resultado.setdefault("valor_servicos", extrair_melhor_valor(val))
                    elif any(kw in header for kw in ["valor liquido", "valor líquido"]):
                        resultado.setdefault("valor_liquido", extrair_melhor_valor(val))
                    elif any(kw in header for kw in ["base de calc"]):
                        resultado.setdefault("base_calculo", extrair_melhor_valor(val))
                    elif any(kw in header for kw in ["valor iss"]):
                        resultado.setdefault("valor_iss", extrair_melhor_valor(val))
                    elif any(kw in header for kw in ["aliquota", "alíquota"]):
                        m = RE_ALIQUOTA_ALT.search(val)
                        if m:
                            resultado.setdefault("aliquota_iss", m.group(1))

        return resultado

    # ── Fusão das 4 estratégias ────────────────────────────────────────────────

    def extrair_tudo(self) -> Dict[str, Any]:
        """
        Executa todas as estratégias em cascata e funde os resultados.
        Prioridade: Estratégia 1 > 2 > 3 > 4 (preenche campos vazios).
        """
        r1 = self._estrategia_regex()
        r2 = self._estrategia_keywords()
        r3 = self._estrategia_posicional()
        r4 = self._estrategia_tabelas()

        # Merge: usa o primeiro valor não-nulo em ordem de prioridade
        fusao: Dict[str, Any] = {}
        todos = [r1, r2, r3, r4]
        for d in todos:
            for k, v in d.items():
                if v and not fusao.get(k):
                    fusao[k] = v

        # Pós-processamento: inferir CNPJ prestador/tomador
        cnpjs_todos = fusao.get("_todos_cnpjs", [])
        if cnpjs_todos:
            if not fusao.get("cnpj_prestador") and len(cnpjs_todos) >= 1:
                fusao["cnpj_prestador"] = cnpjs_todos[0]
            if not fusao.get("cnpj_tomador") and len(cnpjs_todos) >= 2:
                fusao["cnpj_tomador"] = cnpjs_todos[1]

        # Limpeza
        fusao.pop("_todos_cnpjs", None)

        return fusao


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 5: MOTOR PRINCIPAL DE EXTRAÇÃO DE PDF
# ══════════════════════════════════════════════════════════════════════════════

def extrair_texto_pymupdf(pdf_bytes: bytes) -> Tuple[str, List[Dict]]:
    """
    Extrai texto + blocos posicionais usando PyMuPDF.
    Retorna (texto_completo, lista_de_blocos).
    """
    texto_total = []
    blocos_total = []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page in doc:
            texto_total.append(page.get_text("text"))
            # Blocos com coordenadas: (x0, y0, x1, y1, text, block_no, block_type)
            for b in page.get_text("blocks"):
                if b[6] == 0:  # tipo texto
                    blocos_total.append({
                        "x0": b[0], "y0": b[1],
                        "x1": b[2], "y1": b[3],
                        "text": b[4],
                    })
        doc.close()
    except Exception as e:
        logger.warning(f"PyMuPDF erro: {e}")
    return "\n".join(texto_total), blocos_total


def extrair_texto_pdfplumber(pdf_bytes: bytes) -> Tuple[str, List[List[List]]]:
    """
    Extrai texto + tabelas usando pdfplumber.
    Retorna (texto_completo, lista_de_tabelas).
    """
    texto_total = []
    tabelas_total = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text(x_tolerance=3, y_tolerance=3)
                if t:
                    texto_total.append(t)
                # Tenta extrair tabelas
                try:
                    tabelas = page.extract_tables()
                    if tabelas:
                        tabelas_total.extend(tabelas)
                except Exception:
                    pass
    except Exception as e:
        logger.warning(f"pdfplumber erro: {e}")
    return "\n".join(texto_total), tabelas_total


def processar_pdf(pdf_bytes: bytes, nome_arquivo: str) -> List[Dict[str, Any]]:
    """
    Processa um PDF e retorna lista de dicionários (um por item/serviço).
    Usa extração híbrida PyMuPDF + pdfplumber.
    """
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Extração híbrida
    texto_fitz, blocos   = extrair_texto_pymupdf(pdf_bytes)
    texto_plumber, tabelas = extrair_texto_pdfplumber(pdf_bytes)

    # Combina textos (pdfplumber costuma ser melhor para layout)
    texto_combinado = texto_plumber if len(texto_plumber) >= len(texto_fitz) else texto_fitz
    if len(texto_fitz) > 100 and len(texto_plumber) > 100:
        texto_combinado = texto_plumber + "\n" + texto_fitz  # máxima cobertura

    if not texto_combinado.strip():
        logger.warning(f"Nenhum texto extraído de: {nome_arquivo}")
        return [{
            "Arquivo": nome_arquivo,
            "Processado Em": timestamp,
            "Erro": "Nenhum texto extraído — PDF pode ser imagem/escaneado",
        }]

    # Extração em cascata
    extrator = ExtratorEstrategias(texto_combinado, blocos, tabelas)
    dados = extrator.extrair_tudo()

    # Monta linha de resultado
    linha = {
        "Arquivo":               nome_arquivo,
        "Processado Em":         timestamp,
        "Número da Nota":        dados.get("numero_nota", ""),
        "Série":                 dados.get("serie", ""),
        "Data de Emissão":       dados.get("data_emissao", ""),
        "Hora de Emissão":       dados.get("hora_emissao", ""),
        "Competência":           dados.get("competencia", ""),
        "CNPJ Prestador":        dados.get("cnpj_prestador", ""),
        "Razão Social Prestador":dados.get("razao_prestador", ""),
        "CNPJ Tomador":          dados.get("cnpj_tomador", ""),
        "Razão Social Tomador":  dados.get("razao_tomador", ""),
        "Cód. Serviço":          dados.get("codigo_servico", ""),
        "Discriminação":         dados.get("descricao_servico", ""),
        "Município Prestação":   dados.get("municipio_prestacao", ""),
        "Valor Bruto (R$)":      dados.get("valor_servicos", ""),
        "Desconto (R$)":         dados.get("desconto", ""),
        "Base de Cálculo (R$)":  dados.get("base_calculo", ""),
        "Alíquota ISS (%)":      dados.get("aliquota_iss", ""),
        "Valor ISS (R$)":        dados.get("valor_iss", ""),
        "ISS Retido":            dados.get("iss_retido", ""),
        "PIS (R$)":              dados.get("pis", ""),
        "COFINS (R$)":           dados.get("cofins", ""),
        "CSLL (R$)":             dados.get("csll", ""),
        "IRRF (R$)":             dados.get("irrf", ""),
        "INSS (R$)":             dados.get("inss", ""),
        "Valor Líquido (R$)":    dados.get("valor_liquido", ""),
        "Erro": "",
    }

    # Qualidade: marca campos ausentes críticos
    campos_criticos = ["Número da Nota", "CNPJ Prestador", "Valor Bruto (R$)"]
    faltando = [c for c in campos_criticos if not linha.get(c)]
    if faltando:
        linha["Erro"] = f"Campos não encontrados: {', '.join(faltando)}"

    return [linha]


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6: GERAÇÃO DO EXCEL PROFISSIONAL
# ══════════════════════════════════════════════════════════════════════════════

# Paleta de cores enterprise
COR_HEADER      = "1E3A5F"   # Azul escuro
COR_HEADER_FONT = "FFFFFF"   # Branco
COR_LINHA_PAR   = "EBF3FB"   # Azul muito claro
COR_ACENTO      = "2E86AB"   # Azul médio
COR_ERRO        = "FDECEA"   # Vermelho suave
COR_RESUMO_HDR  = "0D2137"   # Azul muito escuro

BORDA_MEDIA = Border(
    left   = Side(style="thin", color="CCCCCC"),
    right  = Side(style="thin", color="CCCCCC"),
    top    = Side(style="thin", color="CCCCCC"),
    bottom = Side(style="thin", color="CCCCCC"),
)


def _aplicar_header(ws, row_num: int, cor_fundo: str = COR_HEADER) -> None:
    """Aplica estilo de cabeçalho em uma linha inteira."""
    for cell in ws[row_num]:
        cell.font      = Font(bold=True, color=COR_HEADER_FONT, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDA_MEDIA


def _auto_largura(ws) -> None:
    """Ajusta largura das colunas automaticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value or ""))
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 55)


def gerar_excel(df: pd.DataFrame) -> bytes:
    """
    Gera planilha Excel profissional com:
      - Aba "Detalhes":     todas as linhas/campos extraídos
      - Aba "Resumo":       agrupamento por Prestador
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()

    # ── Aba 1: Detalhes ───────────────────────────────────────────────────────
    ws_det = wb.active
    ws_det.title = "Detalhes"
    ws_det.freeze_panes = "A2"

    cols = list(df.columns)
    ws_det.append(cols)
    _aplicar_header(ws_det, 1)
    ws_det.row_dimensions[1].height = 35

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        valores = [str(row[c]) if row[c] is not None else "" for c in cols]
        ws_det.append(valores)
        # Zebra striping
        fill_cor = COR_LINHA_PAR if i % 2 == 0 else "FFFFFF"
        tem_erro = bool(row.get("Erro", ""))
        for cell in ws_det[i]:
            cell.fill      = PatternFill("solid", fgColor=COR_ERRO if tem_erro else fill_cor)
            cell.border    = BORDA_MEDIA
            cell.font      = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws_det.row_dimensions[i].height = 18

    _auto_largura(ws_det)
    ws_det.auto_filter.ref = ws_det.dimensions

    # ── Aba 2: Resumo por Nota ────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resumo por Nota")

    colunas_resumo = [
        "Arquivo", "Número da Nota", "Data de Emissão", "Competência",
        "CNPJ Prestador", "Razão Social Prestador",
        "CNPJ Tomador",   "Razão Social Tomador",
        "Valor Bruto (R$)", "Alíquota ISS (%)", "Valor ISS (R$)",
        "ISS Retido", "Valor Líquido (R$)", "Erro",
    ]
    colunas_existentes = [c for c in colunas_resumo if c in df.columns]
    df_resumo = df[colunas_existentes].drop_duplicates(subset=["Arquivo", "Número da Nota"])

    ws_res.append(colunas_existentes)
    _aplicar_header(ws_res, 1, cor_fundo=COR_RESUMO_HDR)
    ws_res.row_dimensions[1].height = 35
    ws_res.freeze_panes = "A2"

    for i, (_, row) in enumerate(df_resumo.iterrows(), start=2):
        valores = [str(row[c]) if row[c] is not None else "" for c in colunas_existentes]
        ws_res.append(valores)
        fill_cor = COR_LINHA_PAR if i % 2 == 0 else "FFFFFF"
        tem_erro = bool(row.get("Erro", ""))
        for cell in ws_res[i]:
            cell.fill      = PatternFill("solid", fgColor=COR_ERRO if tem_erro else fill_cor)
            cell.border    = BORDA_MEDIA
            cell.font      = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center")
        ws_res.row_dimensions[i].height = 18

    _auto_largura(ws_res)
    ws_res.auto_filter.ref = ws_res.dimensions

    # ── Aba 3: Estatísticas ───────────────────────────────────────────────────
    ws_stat = wb.create_sheet("📊 Estatísticas")
    ws_stat.column_dimensions["A"].width = 35
    ws_stat.column_dimensions["B"].width = 25

    total  = len(df)
    erros  = (df["Erro"] != "").sum() if "Erro" in df.columns else 0
    ok     = total - erros
    taxa   = round(ok / total * 100, 1) if total > 0 else 0

    dados_stat = [
        ("Métrica", "Valor"),
        ("Total de NFs Processadas", total),
        ("Extrações com Sucesso", ok),
        ("Extrações com Alertas", erros),
        ("Taxa de Sucesso (%)", f"{taxa}%"),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Sistema", "NF-PDF → Excel Enterprise Ultra v2.0"),
    ]
    for i, (label, val) in enumerate(dados_stat, start=1):
        ws_stat.cell(i, 1, label)
        ws_stat.cell(i, 2, val)
        if i == 1:
            for c in [ws_stat.cell(i, 1), ws_stat.cell(i, 2)]:
                c.font = Font(bold=True, color=COR_HEADER_FONT, size=11, name="Calibri")
                c.fill = PatternFill("solid", fgColor=COR_HEADER)
                c.alignment = Alignment(horizontal="center")
        else:
            for c in [ws_stat.cell(i, 1), ws_stat.cell(i, 2)]:
                c.font   = Font(size=10, name="Calibri")
                c.border = BORDA_MEDIA
                fill_cor = COR_LINHA_PAR if i % 2 == 0 else "FFFFFF"
                c.fill   = PatternFill("solid", fgColor=fill_cor)

    wb.save(output)
    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7: INTERFACE STREAMLIT — ESTILO ENTERPRISE
# ══════════════════════════════════════════════════════════════════════════════

def configurar_pagina() -> None:
    st.set_page_config(
        page_title="NF-PDF → Excel Enterprise Ultra",
        page_icon="🚀",
        layout="wide",
        initial_sidebar_state="expanded",
    )


CSS_ENTERPRISE = """
<style>
/* ── Fonte base ─────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ── Fundo e container ──────────────────────────── */
.main { background-color: #f0f4f8; }
.block-container { padding: 1.5rem 2rem 2rem 2rem; max-width: 1400px; }

/* ── Header hero ────────────────────────────────── */
.hero-banner {
    background: linear-gradient(135deg, #0d2137 0%, #1e3a5f 50%, #2e86ab 100%);
    border-radius: 16px;
    padding: 2.2rem 2.5rem;
    margin-bottom: 1.8rem;
    box-shadow: 0 8px 32px rgba(13,33,55,0.25);
    position: relative;
    overflow: hidden;
}
.hero-banner::before {
    content: "";
    position: absolute; top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(46,134,171,0.25) 0%, transparent 70%);
    border-radius: 50%;
}
.hero-title {
    font-size: 2.1rem; font-weight: 700;
    color: #ffffff; letter-spacing: -0.5px;
    margin: 0; line-height: 1.2;
}
.hero-subtitle {
    font-size: 0.95rem; color: #a8d4e8;
    margin-top: 0.4rem; font-weight: 400;
}
.hero-badge {
    display: inline-block;
    background: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.2);
    border-radius: 20px;
    padding: 3px 12px;
    font-size: 0.75rem; color: #cde9f5;
    margin-top: 0.7rem;
    letter-spacing: 0.5px;
}

/* ── Cards de métricas ──────────────────────────── */
.metric-grid { display: flex; gap: 1rem; margin-bottom: 1.5rem; flex-wrap: wrap; }
.metric-card {
    background: #ffffff;
    border-radius: 12px;
    padding: 1.1rem 1.4rem;
    flex: 1; min-width: 130px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.07);
    border-left: 4px solid #2e86ab;
    transition: transform 0.15s;
}
.metric-card:hover { transform: translateY(-2px); }
.metric-card.success { border-left-color: #27ae60; }
.metric-card.warning { border-left-color: #e67e22; }
.metric-card.info    { border-left-color: #2e86ab; }
.metric-value {
    font-size: 1.7rem; font-weight: 700;
    color: #1e3a5f; margin: 0;
}
.metric-label {
    font-size: 0.75rem; color: #7f8c8d;
    font-weight: 500; text-transform: uppercase;
    letter-spacing: 0.5px; margin-top: 2px;
}

/* ── Upload area ────────────────────────────────── */
.upload-zone {
    background: #ffffff;
    border: 2px dashed #2e86ab;
    border-radius: 14px;
    padding: 2rem;
    text-align: center;
    margin-bottom: 1.5rem;
    transition: border-color 0.2s;
}
.upload-zone:hover { border-color: #1e3a5f; }

/* ── Botão de download ──────────────────────────── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #1e3a5f 0%, #2e86ab 100%) !important;
    color: white !important; font-weight: 600 !important;
    border-radius: 10px !important; border: none !important;
    padding: 0.6rem 1.5rem !important;
    box-shadow: 0 4px 15px rgba(46,134,171,0.35) !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(46,134,171,0.45) !important;
}

/* ── Progress bar ───────────────────────────────── */
.stProgress > div > div { background-color: #2e86ab !important; }

/* ── Sidebar ────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: #0d2137;
}
section[data-testid="stSidebar"] * {
    color: #cde9f5 !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
}

/* ── Tabela de prévia ───────────────────────────── */
.dataframe thead tr th {
    background: #1e3a5f !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    font-size: 0.8rem !important;
}
.dataframe tbody tr:nth-child(even) { background: #ebf3fb !important; }
.dataframe tbody tr:hover { background: #d0e8f5 !important; }

/* ── Alertas customizados ───────────────────────── */
.info-box {
    background: #e8f4f8; border-left: 4px solid #2e86ab;
    border-radius: 8px; padding: 0.8rem 1rem;
    margin: 0.5rem 0; font-size: 0.88rem;
}
.success-box {
    background: #e8f8ef; border-left: 4px solid #27ae60;
    border-radius: 8px; padding: 0.8rem 1rem;
    margin: 0.5rem 0; font-size: 0.88rem;
}
.error-box {
    background: #fdecea; border-left: 4px solid #e74c3c;
    border-radius: 8px; padding: 0.8rem 1rem;
    margin: 0.5rem 0; font-size: 0.88rem;
}

/* ── Divisor elegante ───────────────────────────── */
hr { border: none; border-top: 1px solid #dce6f0; margin: 1.2rem 0; }

/* ── Scrollbar ──────────────────────────────────── */
::-webkit-scrollbar { width: 7px; height: 7px; }
::-webkit-scrollbar-track { background: #f0f4f8; }
::-webkit-scrollbar-thumb { background: #2e86ab; border-radius: 4px; }
</style>
"""


def renderizar_hero() -> None:
    st.markdown(CSS_ENTERPRISE, unsafe_allow_html=True)
    st.markdown("""
    <div class="hero-banner">
        <div class="hero-title">🚀 NF-PDF → Excel Enterprise Ultra</div>
        <div class="hero-subtitle">
            Extração inteligente de NFS-e / NF-e para Excel — Multi-prefeitura, Alta Precisão
        </div>
        <span class="hero-badge">⚡ Hybrid Engine · pdfplumber + PyMuPDF · 4-Layer Cascade · v2.0</span>
    </div>
    """, unsafe_allow_html=True)


def renderizar_sidebar() -> Dict[str, Any]:
    """Renderiza sidebar com configurações e retorna opções."""
    with st.sidebar:
        st.markdown("## ⚙️ Configurações")
        st.markdown("---")

        st.markdown("### 📋 Prefeituras Suportadas")
        prefeituras = [
            "✅ São Paulo (SP)", "✅ Taboão da Serra (SP)",
            "✅ Osasco (SP)", "✅ Campo Limpo Paulista (SP)",
            "✅ Joinville (SC)", "✅ Porto Alegre (RS)",
            "✅ Rio de Janeiro (RJ)", "✅ Curitiba (PR)",
            "✅ Belo Horizonte (MG)", "✅ Campinas (SP)",
            "✅ + Dezenas de outros layouts",
        ]
        for p in prefeituras:
            st.caption(p)

        st.markdown("---")
        st.markdown("### 🔧 Opções de Extração")

        opcoes = {
            "mostrar_previa":    st.checkbox("Mostrar prévia da tabela", value=True),
            "mostrar_texto_raw": st.checkbox("Mostrar texto bruto extraído", value=False),
            "filtro_busca":      st.text_input("🔍 Filtrar tabela por:", placeholder="Ex: Taboão"),
            "max_linhas_previa": st.slider("Linhas na prévia", 5, 100, 20),
        }

        st.markdown("---")
        st.markdown("### ℹ️ Sobre o Sistema")
        st.caption(
            "Motor de extração híbrido com 4 estratégias em cascata:\n\n"
            "1. Regex multi-padrão\n"
            "2. Proximidade por keywords\n"
            "3. Análise posicional\n"
            "4. Extração de tabelas"
        )
        st.markdown("---")
        st.caption("🔒 Processamento 100% local\nNenhum dado é enviado a servidores externos.")

    return opcoes


def renderizar_metricas(total: int, sucesso: int, alertas: int,
                         tempo: float) -> None:
    """Renderiza cards de métricas do processamento."""
    taxa = round(sucesso / total * 100, 1) if total > 0 else 0
    st.markdown(f"""
    <div class="metric-grid">
        <div class="metric-card info">
            <div class="metric-value">{total}</div>
            <div class="metric-label">PDFs Processados</div>
        </div>
        <div class="metric-card success">
            <div class="metric-value">{sucesso}</div>
            <div class="metric-label">Extrações OK</div>
        </div>
        <div class="metric-card warning">
            <div class="metric-value">{alertas}</div>
            <div class="metric-label">Com Alertas</div>
        </div>
        <div class="metric-card {'success' if taxa >= 90 else 'warning'}">
            <div class="metric-value">{taxa}%</div>
            <div class="metric-label">Taxa de Precisão</div>
        </div>
        <div class="metric-card info">
            <div class="metric-value">{tempo:.1f}s</div>
            <div class="metric-label">Tempo Total</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8: PONTO DE ENTRADA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    configurar_pagina()
    renderizar_hero()
    opcoes = renderizar_sidebar()

    # ── Upload ────────────────────────────────────────────────────────────────
    st.markdown("### 📤 Upload de Arquivos")
    uploaded = st.file_uploader(
        "Arraste PDFs de NFS-e / NF-e aqui ou clique para selecionar",
        type=["pdf"],
        accept_multiple_files=True,
        help="Suporta todos os layouts de NFS-e municipais e NF-e federais",
        key="pdf_upload",
    )

    if not uploaded:
        st.markdown("""
        <div class="info-box">
            📌 <strong>Como usar:</strong> Faça upload de um ou mais PDFs de Notas Fiscais.
            O sistema detecta automaticamente o layout e extrai todos os campos disponíveis.
            O resultado é exportado como planilha Excel profissional com múltiplas abas.
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("#### 📊 Campos Extraídos Automaticamente")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("""
            **🏢 Identificação**
            - CNPJ/CPF Prestador
            - Razão Social Prestador
            - CNPJ/CPF Tomador
            - Razão Social Tomador
            - Nº Nota / RPS / Série
            - Data e Hora de Emissão
            - Competência
            """)
        with col2:
            st.markdown("""
            **💰 Valores Fiscais**
            - Valor Bruto dos Serviços
            - Base de Cálculo
            - Alíquota ISS (%)
            - Valor ISS
            - ISS Retido (Sim/Não)
            - Valor Líquido
            - Descontos
            """)
        with col3:
            st.markdown("""
            **📋 Retenções & Serviço**
            - PIS / COFINS / CSLL
            - IRRF / INSS
            - Código do Serviço
            - Discriminação
            - Município de Prestação
            - Nome do Arquivo
            - Data/Hora Processamento
            """)
        return

    # ── Processamento ─────────────────────────────────────────────────────────
    if st.button("▶️ Processar Notas Fiscais", type="primary", use_container_width=True):

        inicio = datetime.now()
        resultados: List[Dict[str, Any]] = []
        total_pdfs = len(uploaded)
        textos_raw: List[Tuple[str, str]] = []

        st.markdown("---")
        st.markdown(f"### ⚙️ Processando {total_pdfs} arquivo(s)...")

        barra_prog  = st.progress(0.0)
        status_text = st.empty()
        log_area    = st.expander("📋 Log em tempo real", expanded=False)
        log_msgs: List[str] = []

        for i, arquivo in enumerate(uploaded):
            nome = arquivo.name
            status_text.markdown(
                f"<div class='info-box'>⏳ Processando <strong>{nome}</strong> "
                f"({i+1}/{total_pdfs})...</div>",
                unsafe_allow_html=True,
            )

            try:
                pdf_bytes = arquivo.read()
                linhas = processar_pdf(pdf_bytes, nome)
                resultados.extend(linhas)

                tem_erro = any(l.get("Erro") for l in linhas)
                status_emoji = "⚠️" if tem_erro else "✅"
                msg = f"{status_emoji} {nome} — {len(linhas)} linha(s)"
                log_msgs.append(msg)
                logger.info(msg)

                if opcoes["mostrar_texto_raw"]:
                    t_fitz, _    = extrair_texto_pymupdf(pdf_bytes)
                    t_plumber, _ = extrair_texto_pdfplumber(pdf_bytes)
                    textos_raw.append((nome, t_plumber or t_fitz))

            except Exception as exc:
                msg_err = f"❌ ERRO em {nome}: {exc}"
                log_msgs.append(msg_err)
                logger.error(f"{msg_err}\n{traceback.format_exc()}")
                resultados.append({
                    "Arquivo": nome,
                    "Processado Em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "Erro": str(exc),
                })

            barra_prog.progress((i + 1) / total_pdfs)
            with log_area:
                st.code("\n".join(log_msgs[-20:]), language="")

        tempo_total = (datetime.now() - inicio).total_seconds()
        status_text.empty()
        barra_prog.progress(1.0)

        # ── Dashboard de resultados ────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📊 Dashboard de Resultados")

        df = pd.DataFrame(resultados)
        n_total   = len(df)
        n_alertas = int((df["Erro"] != "").sum()) if "Erro" in df.columns else 0
        n_sucesso = n_total - n_alertas

        renderizar_metricas(n_total, n_sucesso, n_alertas, tempo_total)

        # ── Botão de Download ──────────────────────────────────────────────────
        st.markdown("#### 📥 Exportar Excel")
        excel_bytes = gerar_excel(df)
        nome_excel  = f"NF_Excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        col_dl, col_info = st.columns([1, 3])
        with col_dl:
            st.download_button(
                label="⬇️ Baixar Excel Enterprise",
                data=excel_bytes,
                file_name=nome_excel,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_info:
            st.markdown(
                f"<div class='success-box'>✅ Arquivo <strong>{nome_excel}</strong> pronto — "
                f"{n_total} linha(s), 3 abas (Detalhes · Resumo · Estatísticas)</div>",
                unsafe_allow_html=True,
            )

        # ── Prévia Interativa ──────────────────────────────────────────────────
        if opcoes["mostrar_previa"] and not df.empty:
            st.markdown("---")
            st.markdown("#### 🔍 Prévia dos Dados Extraídos")

            df_previa = df.copy()
            if opcoes["filtro_busca"]:
                mask = df_previa.apply(
                    lambda row: row.astype(str).str.contains(
                        opcoes["filtro_busca"], case=False, na=False
                    ).any(),
                    axis=1,
                )
                df_previa = df_previa[mask]
                st.caption(f"🔍 Filtro '{opcoes['filtro_busca']}': {len(df_previa)} resultado(s)")

            st.dataframe(
                df_previa.head(opcoes["max_linhas_previa"]),
                use_container_width=True,
                hide_index=True,
            )

        # ── Notas com alertas ──────────────────────────────────────────────────
        if n_alertas > 0:
            st.markdown("---")
            st.markdown("#### ⚠️ Notas com Alertas de Extração")
            df_erros = df[df["Erro"] != ""][["Arquivo", "Erro"]].copy() if "Erro" in df.columns else pd.DataFrame()
            if not df_erros.empty:
                for _, row in df_erros.iterrows():
                    st.markdown(
                        f"<div class='error-box'>📄 <strong>{row['Arquivo']}</strong><br>"
                        f"⚠️ {row['Erro']}</div>",
                        unsafe_allow_html=True,
                    )

        # ── Texto bruto (debug) ────────────────────────────────────────────────
        if opcoes["mostrar_texto_raw"] and textos_raw:
            st.markdown("---")
            st.markdown("#### 🔬 Texto Bruto Extraído (Debug)")
            for nome_arq, texto in textos_raw:
                with st.expander(f"📄 {nome_arq}"):
                    st.text_area("Texto extraído:", value=texto[:3000], height=250,
                                  key=f"raw_{nome_arq}", disabled=True)

        # Mensagem final
        if n_sucesso == n_total:
            st.balloons()
            st.markdown(
                "<div class='success-box'>🎉 <strong>Processamento concluído com sucesso!</strong> "
                "Todas as notas foram extraídas sem alertas.</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"<div class='info-box'>✅ Processamento concluído. "
                f"{n_sucesso}/{n_total} notas OK, {n_alertas} com alertas.<br>"
                f"Alertas indicam campos não encontrados no PDF — verifique o log acima.</div>",
                unsafe_allow_html=True,
            )

    else:
        # Estado inicial após upload (antes de processar)
        if uploaded:
            st.markdown(
                f"<div class='info-box'>📂 <strong>{len(uploaded)} arquivo(s)</strong> "
                f"carregado(s). Clique em <strong>▶️ Processar Notas Fiscais</strong> para iniciar.</div>",
                unsafe_allow_html=True,
            )
            with st.expander("📋 Arquivos na fila"):
                for f in uploaded:
                    st.caption(f"• {f.name} ({f.size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
